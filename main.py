from tkinter import *
import train
from PIL import Image,ImageTk
import face_recognition
from openpyxl import Workbook
from openpyxl import load_workbook
import itertools

def open_train():
    train.train()
def open_detect():
    face_recognition.recognition(inputform.get())

def get_name():
    wb=Workbook()
    wb=load_workbook('Attendance.xlsx')
    ws=wb.active
    col_a=ws['A']
    col_b=ws['B']
    lista=''
    for (cell,cell2) in zip(col_a,col_b):
      lista=f'{lista +str(cell.value)+" "+str(cell2.value)}\n'
    labeldisplay.config(text=lista)

root=Tk()
root.geometry("500x800")
root.title('Attendance System')
image = Image.open("GUI/register.jpg")
image=image.resize((580,600), Image.ANTIALIAS)
photo = ImageTk.PhotoImage(image)
labela=Label(root,text="ATTENDANCE SYSTEM")
labela.pack(pady="20")
labelentry=Label(root,text="enter file path")
inputform=Entry(root,text="enter data")
inputform.pack(pady="20")

buttontrain=Button(root,text="train model",command=open_train)
buttontrain.pack(pady=20)
buttondetect=Button(root,text="detect model",command=open_detect)
buttondetect.pack(pady=20)

buttondisplay=Button(root,text="display attendance",command=get_name)
buttondisplay.pack(pady=20)
labeldisplay=Label(root,text="")
labeldisplay.pack(pady=20)

root.mainloop()